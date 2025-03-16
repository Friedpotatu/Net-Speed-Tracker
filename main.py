import speedtest
import pandas as pd
import psutil
import subprocess
from datetime import datetime
import os
from openpyxl import load_workbook

# Nombre del archivo Excel
file_name = "network_speed.xlsx"

# Función para medir velocidad
def medir_velocidad():
    st = speedtest.Speedtest(secure=True)
    st.get_best_server()
    download_speed = round(st.download() / 1_000_000, 2)  # Convertir a Mbps y redondear
    upload_speed = round(st.upload() / 1_000_000, 2)  # Convertir a Mbps y redondear
    ping = round(st.results.ping, 2)  # Redondear a 2 decimales
    return download_speed, upload_speed, ping

# Función para obtener la red conectada
def obtener_red():
    try:
        # Para Windows
        if os.name == "nt":
            result = subprocess.run(["netsh", "wlan", "show", "interfaces"], capture_output=True, text=True)
            for line in result.stdout.split("\n"):
                if "SSID" in line:
                    return line.split(":")[1].strip()
        # Para macOS y Linux
        else:
            result = subprocess.run(["iwgetid", "-r"], capture_output=True, text=True)
            return result.stdout.strip()
    except Exception:
        return "Desconocida"

# Función para obtener la interfaz activa
def obtener_interfaz():
    interfaces = psutil.net_if_addrs()
    for interface in interfaces:
        if "Wi-Fi" in interface or "wlan" in interface.lower():
            return "Wi-Fi"
        elif "eth" in interface.lower():
            return "Ethernet"
    return "Desconocida"

# Obtener datos
download, upload, ping = medir_velocidad()
fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
red_conectada = obtener_red()
interfaz = obtener_interfaz()

# Crear DataFrame con los datos
data = {
    "Fecha y Hora": [fecha_hora],
    "Red Conectada (SSID)": [red_conectada],
    "Interfaz": [interfaz],
    "Velocidad Descarga (Mbps)": [download],
    "Velocidad Subida (Mbps)": [upload],
    "Ping (ms)": [ping],
}
df = pd.DataFrame(data)

# Guardar en Excel (agregando si ya existe)
if os.path.exists(file_name):
    with pd.ExcelWriter(file_name, mode="a", if_sheet_exists="overlay", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Datos", index=False, header=False, startrow=writer.sheets["Datos"].max_row)
else:
    df.to_excel(file_name, sheet_name="Datos", index=False)

# Ajustar los anchos de columna automáticamente
wb = load_workbook(file_name)
ws = wb["Datos"]

for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2  # Ajuste extra

wb.save(file_name)

print(f"✅ Medición guardada en {file_name}: {red_conectada} - {interfaz}")
